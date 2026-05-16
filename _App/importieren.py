#!/usr/bin/env python3
"""
Imkerei Buchhaltung – Import aus GitHub
Holt data.json von GitHub und schreibt neue Einträge in die Excel-Datei.
"""
import json, sys, datetime, urllib.request
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).parent
ROOT     = BASE.parent
EXCEL    = ROOT / "Imkerei_Buchhaltung.xlsx"
IMPORTED = ROOT / ".importiert_ids.json"

GH_CONTENTS = "https://api.github.com/repos/hanging-pawn/imkerei/contents/"

OFF_WHITE = "F8FAFC"; WHITE = "FFFFFF"
THIN      = Side(style="thin", color="E2E8F0")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AMBER_SOFT= "FEF3C7"

def style(cell, row, bold=False, align="left", fmt=None, color="0F172A", highlight=False):
    bg = AMBER_SOFT if highlight else (OFF_WHITE if row % 2 == 0 else WHITE)
    cell.font      = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align=="left" else 0)
    if fmt: cell.number_format = fmt

def next_row(ws, start=4):
    for r in range(start, 1100):
        if ws.cell(r, 1).value is None: return r
    return 1100

def load_ids():
    if IMPORTED.exists():
        try: return set(json.loads(IMPORTED.read_text(encoding="utf-8")))
        except: pass
    return set()

def save_ids(ids):
    IMPORTED.write_text(json.dumps(sorted(ids)), encoding="utf-8")

def content_key_ein(e):
    return f"{e.get('datum')}|{e.get('produkt')}|{e.get('menge')}|{e.get('preis')}|{e.get('kanal')}"

def content_key_aus(e):
    return f"{e.get('datum')}|{e.get('kategorie')}|{e.get('beschreibung')}|{e.get('betrag')}"

def existing_keys_ein(ws):
    keys = set()
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r,1).value
        if d is None: continue
        ds = d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
        keys.add(f"{ds}|{ws.cell(r,3).value}|{ws.cell(r,4).value}|{ws.cell(r,5).value}|{ws.cell(r,7).value}")
    return keys

def existing_keys_aus(ws):
    keys = set()
    for r in range(4, ws.max_row + 1):
        d = ws.cell(r,1).value
        if d is None: continue
        ds = d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
        keys.add(f"{ds}|{ws.cell(r,3).value}|{ws.cell(r,4).value}|{ws.cell(r,5).value}")
    return keys

def fetch_entries():
    print("  → Suche Datendateien auf GitHub…", end=" ", flush=True)
    try:
        req = urllib.request.Request(GH_CONTENTS, headers={"Cache-Control": "no-cache", "User-Agent": "imkerei-import"})
        with urllib.request.urlopen(req, timeout=15) as r:
            files = json.loads(r.read().decode("utf-8"))
        data_files = [f for f in files if f["name"].startswith("data") and f["name"].endswith(".json")]
        print(f"{len(data_files)} Datei(en) gefunden")
    except Exception as e:
        print(f"FEHLER: {e}")
        return None

    all_entries = []
    seen_ids = set()
    for f in data_files:
        try:
            print(f"  → Lese {f['name']}…", end=" ", flush=True)
            req2 = urllib.request.Request(f["download_url"], headers={"Cache-Control": "no-cache"})
            with urllib.request.urlopen(req2, timeout=15) as r:
                data = json.loads(r.read().decode("utf-8"))
            entries = data.get("entries", [])
            neu = 0
            for e in entries:
                eid = e.get("id")
                if eid and eid in seen_ids: continue
                if eid: seen_ids.add(eid)
                all_entries.append(e)
                neu += 1
            print(f"{neu} Einträge")
        except Exception as e:
            print(f"Warnung: {e}")

    print(f"  → Total: {len(all_entries)} Einträge")
    return all_entries

def main():
    print("\n" + "="*54)
    print("  Imkerei Buchhaltung — Daten importieren")
    print("="*54 + "\n")

    if not EXCEL.exists():
        print(f"  FEHLER: Excel-Datei nicht gefunden:\n  {EXCEL}\n")
        input("  Enter zum Beenden..."); sys.exit(1)

    all_entries = fetch_entries()
    if all_entries is None:
        print("\n  Kein Internet oder GitHub nicht erreichbar.")
        input("  Enter zum Beenden..."); return
    if not all_entries:
        print("  Keine Einträge gefunden.\n")
        input("  Enter zum Beenden..."); return

    imported_ids = load_ids()
    wb = load_workbook(EXCEL)
    ws_ein = wb["Einnahmen"]
    ws_aus = wb["Ausgaben"]
    ex_ein = existing_keys_ein(ws_ein)
    ex_aus = existing_keys_aus(ws_aus)

    neu_ein = neu_aus = skip = 0

    for entry in sorted(all_entries, key=lambda x: x.get("datum","")):
        eid = str(entry.get("id",""))
        typ = entry.get("type","")
        if eid and eid in imported_ids: skip += 1; continue

        if typ == "ein":
            ck = content_key_ein(entry)
            if ck in ex_ein:
                if eid: imported_ids.add(eid); skip += 1; continue
            row = next_row(ws_ein)
            try: d = datetime.datetime.strptime(entry["datum"], "%Y-%m-%d")
            except: d = datetime.datetime.today()
            ws_ein.cell(row,1).value = d;  style(ws_ein.cell(row,1), row, fmt="DD.MM.YYYY", align="center")
            ws_ein.cell(row,3).value = entry.get("produkt",""); style(ws_ein.cell(row,3), row)
            ws_ein.cell(row,4).value = float(entry.get("menge",0)); style(ws_ein.cell(row,4), row, fmt='#,##0.00', align="right")
            ws_ein.cell(row,5).value = float(entry.get("preis",0)); style(ws_ein.cell(row,5), row, fmt='#,##0.00 "CHF"', align="right")
            ws_ein.cell(row,7).value = entry.get("kanal",""); style(ws_ein.cell(row,7), row)
            ws_ein.cell(row,8).value = entry.get("notiz",""); style(ws_ein.cell(row,8), row, color="64748B")
            ex_ein.add(ck)
            if eid: imported_ids.add(eid)
            g = float(entry.get("gesamt", float(entry.get("menge",0))*float(entry.get("preis",0))))
            print(f"  + {entry.get('datum')}  {entry.get('produkt',''):<22}  {g:>8.2f} CHF")
            neu_ein += 1

        elif typ == "aus":
            ck = content_key_aus(entry)
            if ck in ex_aus:
                if eid: imported_ids.add(eid); skip += 1; continue
            row = next_row(ws_aus)
            try: d = datetime.datetime.strptime(entry["datum"], "%Y-%m-%d")
            except: d = datetime.datetime.today()
            ws_aus.cell(row,1).value = d;  style(ws_aus.cell(row,1), row, fmt="DD.MM.YYYY", align="center")
            ws_aus.cell(row,3).value = entry.get("kategorie",""); style(ws_aus.cell(row,3), row)
            ws_aus.cell(row,4).value = entry.get("beschreibung",""); style(ws_aus.cell(row,4), row)
            ws_aus.cell(row,5).value = float(entry.get("betrag",0)); style(ws_aus.cell(row,5), row, fmt='#,##0.00 "CHF"', align="right", bold=True, highlight=True)
            ws_aus.cell(row,6).value = entry.get("notiz",""); style(ws_aus.cell(row,6), row, color="64748B")
            ex_aus.add(ck)
            if eid: imported_ids.add(eid)
            print(f"  - {entry.get('datum')}  {entry.get('kategorie',''):<22}  {float(entry.get('betrag',0)):>8.2f} CHF")
            neu_aus += 1

    print("\n" + "-"*54)
    if neu_ein + neu_aus == 0:
        print("  Alles aktuell — keine neuen Einträge.")
    else:
        wb.save(EXCEL)
        save_ids(imported_ids)
        print(f"  Importiert: {neu_ein} Einnahme(n)  |  {neu_aus} Ausgabe(n)")
    if skip: print(f"  Übersprungen: {skip} (bereits vorhanden)")
    print("-"*54 + "\n")

    input("  Enter zum Beenden...")

if __name__ == "__main__":
    main()
